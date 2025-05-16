import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def searchnext():
    global now_line
    global lines
    if now_line < len(lines):
        while now_line < len(lines) and len(lines[now_line]) == 0:
            now_line += 1

def insertexcel(lines):
    now_line = 0;

    be_line = 0
    while be_line < len(lines) and len(lines[be_line]) == 0:
        be_line += 1
    #获取表名
    if be_line >= len(lines):
        return
    table_n = re.findall(r'create table\s+(\w+\.\w+)', lines[be_line])[0]
    table_space = table_n.split(".")[0] #表所属空间
    table_name = table_n.split(".")[1] #表名字

    list = [] #表项列表
    pr_l = {} #主键字典
    ex_l = {} #表项注释

    #
    #提取表元素
    for i in range(be_line+1 , len(lines)):
        if ');' in lines[i]:
            now_line = i+1
            break
        elements = lines[i].split()
        data = {}
        data['name'] = elements[0]
        typelis = elements[1].split(',')
        if(len(typelis) >= 2 and typelis[1] != ''):
            data['type'] = typelis[0]+','+ typelis[1]
        else:
            data['type'] = typelis[0]
        # 有约束条件
        data['default'] = 'N' #默认没有约束
        data['allownull'] = 'Y'  # 不设约束
        if(len(elements) > 2):
            # 只有非空约束
            if elements[2] == 'not':
                data['allownull'] = 'N'
                data['default'] = 'N'
            # 有默认值设置
            if elements[2] == 'default':
                data['default'] = elements[3].split(',')[0]
                if len(elements) > 4 and elements[4] == 'not':# 既有默认值约束又有非空约束
                    data['allownull'] = 'N'
        list.append(data)

    #print(list)
    while now_line < len(lines) and len(lines[now_line]) != 0:
        #searchnext();
        if 'primary key' in lines[now_line]:
            matches = re.findall(r'primary\s+key\s*\(\s*(.*?)\s*\)', lines[now_line])
            if matches:
                # print(matches)
                pr_val = matches[0].split(',')
                for i in pr_val:
                    pr_l[i] = 'Y'
            #print(pr_l)
        if 'comment' in lines[now_line]:
            print(lines[now_line])
            fin = re.findall(r'comment on column\s+(\w+)', lines[now_line])
            if len(fin) > 0:
                name = re.findall(r'comment on column\s+(\w+)', lines[now_line])[0]
                discrip = re.findall(r'is \'(.*?)\'', lines[now_line])[0]
                ex_l[name] = discrip
        now_line+=1

    data = []
    for i in list:
        now_da = {}
        now_da['字段'] = i['name']
        now_da['字段类型'] = i['type']
        now_da['是否必填'] = '必填' if i['allownull'] == 'N' else '非必填'
        now_da['默认值'] = i['default'] if i['default'] != 'N' else ''
        now_da['是否主键'] = '主键' if i['name'] in pr_l else ''
        now_da['注释'] = ex_l[i['name']] if i['name'] in ex_l else ''
        data.append(now_da)
    file_name = 'output.xlsx'
    cloumns = ['字段', '字段类型', '是否必填', '默认值', '是否主键', '注释']
    df = pd.DataFrame(data, columns=cloumns)
    with pd.ExcelWriter(file_name, mode='a', engine="openpyxl" ,if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name=table_name)
    wb = load_workbook(file_name)
    ws = wb[table_name]  # 获取指定工作表

    # 定义边框样式（细线边框）
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 遍历所有单元格并应用边框
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    ws.column_dimensions['A'].width = 25  # 设置列A的宽度
    ws.column_dimensions['B'].width = 16  # 设置列B的宽度
    ws.column_dimensions['C'].width = 10  # 设置列C的宽度
    ws.column_dimensions['D'].width = 10  # 设置列D的宽度
    ws.column_dimensions['E'].width = 10  # 设置列E的宽度
    ws.column_dimensions['F'].width = 80  # 设置列F的宽度
    # 保存修改后的文件
    wb.save(file_name)